import openpyxl
path = 'C:/Users/user/Desktop/xyz.xlsx'

#load workbook
work_book = openpyxl.load_workbook(path)

#open active sheet
#type-1
sheet = work_book.active
# type-2
# sheet1 = work_book.get_sheet_by_name('1569327240889b8ghijcvBTjqG6o9')

#count no. of rows
row_count= sheet.max_row

# count no. of cols
col_count = sheet.max_column

print (row_count)
print(col_count)

for r in range(1,row_count+1):
    for c in range(2,col_count+1):
        print(sheet.cell(row=r, column=c).value, end = '            ')
    print()